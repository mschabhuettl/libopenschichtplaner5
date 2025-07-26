# libopenschichtplaner5/src/libopenschichtplaner5/export.py
"""
Export module for Schichtplaner5 data.
Supports multiple export formats including Excel, CSV, JSON, and HTML reports.
"""

import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime, date
from io import StringIO, BytesIO
import html

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportFormat:
    """Supported export formats."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    HTML = "html"
    MARKDOWN = "markdown"


class DataExporter:
    """Export Schichtplaner5 data to various formats."""

    def __init__(self):
        self.default_datetime_format = "%Y-%m-%d %H:%M:%S"
        self.default_date_format = "%Y-%m-%d"

    def export(self, data: Union[List[Dict], Dict],
               format: str,
               output_path: Optional[Path] = None,
               **kwargs) -> Union[str, bytes]:
        """
        Export data to specified format.

        Args:
            data: Data to export (list of dicts or single dict)
            format: Export format (csv, json, excel, html, markdown)
            output_path: Optional path to save the export
            **kwargs: Format-specific options

        Returns:
            Exported data as string or bytes
        """
        if format == ExportFormat.CSV:
            result = self.to_csv(data, **kwargs)
        elif format == ExportFormat.JSON:
            result = self.to_json(data, **kwargs)
        elif format == ExportFormat.EXCEL:
            result = self.to_excel(data, **kwargs)
        elif format == ExportFormat.HTML:
            result = self.to_html(data, **kwargs)
        elif format == ExportFormat.MARKDOWN:
            result = self.to_markdown(data, **kwargs)
        else:
            raise ValueError(f"Unsupported format: {format}")

        if output_path:
            if isinstance(result, bytes):
                output_path.write_bytes(result)
            else:
                output_path.write_text(result, encoding='utf-8')

        return result

    def to_csv(self, data: Union[List[Dict], Dict],
               delimiter: str = ",",
               include_headers: bool = True) -> str:
        """Export data to CSV format."""
        output = StringIO()

        # Ensure we have a list
        if isinstance(data, dict):
            data = [data]

        if not data:
            return ""

        # Get all unique fields
        fields = set()
        for record in data:
            fields.update(self._flatten_dict(record).keys())
        fields = sorted(list(fields))

        writer = csv.DictWriter(output, fieldnames=fields, delimiter=delimiter)

        if include_headers:
            writer.writeheader()

        for record in data:
            flat_record = self._flatten_dict(record)
            writer.writerow(flat_record)

        return output.getvalue()

    def to_json(self, data: Union[List[Dict], Dict],
                indent: int = 2,
                ensure_ascii: bool = False) -> str:
        """Export data to JSON format."""

        def json_serializer(obj):
            if isinstance(obj, (date, datetime)):
                return obj.isoformat()
            elif hasattr(obj, '__dict__'):
                return obj.__dict__
            return str(obj)

        return json.dumps(data, indent=indent, ensure_ascii=ensure_ascii,
                          default=json_serializer)

    def to_excel(self, data: Union[List[Dict], Dict],
                 sheet_name: str = "Data",
                 include_headers: bool = True,
                 auto_filter: bool = True,
                 freeze_panes: bool = True) -> bytes:
        """Export data to Excel format."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Ensure we have a list
        if isinstance(data, dict):
            data = [data]

        if not data:
            return bytes()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Get all unique fields
        fields = set()
        for record in data:
            fields.update(self._flatten_dict(record).keys())
        fields = sorted(list(fields))

        # Write headers
        if include_headers:
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = field
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, record in enumerate(data, 2 if include_headers else 1):
            flat_record = self._flatten_dict(record)
            for col_idx, field in enumerate(fields, 1):
                value = flat_record.get(field, "")
                if isinstance(value, (date, datetime)):
                    value = value.strftime(
                        self.default_date_format if isinstance(value, date) else self.default_datetime_format)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for col_idx, field in enumerate(fields, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(field))

            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Add auto filter
        if auto_filter and include_headers:
            ws.auto_filter.ref = ws.dimensions

        # Freeze top row
        if freeze_panes and include_headers:
            ws.freeze_panes = ws['A2']

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def to_html(self, data: Union[List[Dict], Dict],
                title: str = "Data Export",
                include_style: bool = True,
                table_class: str = "data-table") -> str:
        """Export data to HTML table format."""
        # Ensure we have a list
        if isinstance(data, dict):
            data = [data]

        if not data:
            return "<p>No data to display</p>"

        # Get all unique fields
        fields = set()
        for record in data:
            fields.update(self._flatten_dict(record).keys())
        fields = sorted(list(fields))

        # Build HTML
        html_parts = []

        # Start HTML
        html_parts.append('<!DOCTYPE html>')
        html_parts.append('<html>')
        html_parts.append('<head>')
        html_parts.append(f'<title>{html.escape(title)}</title>')
        html_parts.append('<meta charset="utf-8">')

        # Add style if requested
        if include_style:
            html_parts.append('<style>')
            html_parts.append(self._get_default_css())
            html_parts.append('</style>')

        html_parts.append('</head>')
        html_parts.append('<body>')
        html_parts.append(f'<h1>{html.escape(title)}</h1>')

        # Build table
        html_parts.append(f'<table class="{table_class}">')

        # Headers
        html_parts.append('<thead>')
        html_parts.append('<tr>')
        for field in fields:
            html_parts.append(f'<th>{html.escape(field)}</th>')
        html_parts.append('</tr>')
        html_parts.append('</thead>')

        # Data
        html_parts.append('<tbody>')
        for record in data:
            flat_record = self._flatten_dict(record)
            html_parts.append('<tr>')
            for field in fields:
                value = flat_record.get(field, "")
                if isinstance(value, (date, datetime)):
                    value = value.strftime(
                        self.default_date_format if isinstance(value, date) else self.default_datetime_format)
                html_parts.append(f'<td>{html.escape(str(value))}</td>')
            html_parts.append('</tr>')
        html_parts.append('</tbody>')

        html_parts.append('</table>')
        html_parts.append('</body>')
        html_parts.append('</html>')

        return '\n'.join(html_parts)

    def to_markdown(self, data: Union[List[Dict], Dict],
                    title: Optional[str] = None) -> str:
        """Export data to Markdown table format."""
        # Ensure we have a list
        if isinstance(data, dict):
            data = [data]

        if not data:
            return "No data to display"

        # Get all unique fields
        fields = set()
        for record in data:
            fields.update(self._flatten_dict(record).keys())
        fields = sorted(list(fields))

        # Build markdown
        lines = []

        if title:
            lines.append(f"# {title}")
            lines.append("")

        # Headers
        lines.append("| " + " | ".join(fields) + " |")
        lines.append("| " + " | ".join(["-" * max(len(f), 3) for f in fields]) + " |")

        # Data
        for record in data:
            flat_record = self._flatten_dict(record)
            values = []
            for field in fields:
                value = flat_record.get(field, "")
                if isinstance(value, (date, datetime)):
                    value = value.strftime(
                        self.default_date_format if isinstance(value, date) else self.default_datetime_format)
                # Escape pipe characters in values
                value = str(value).replace("|", "\\|")
                values.append(value)
            lines.append("| " + " | ".join(values) + " |")

        return '\n'.join(lines)

    def _flatten_dict(self, d: Dict[str, Any], parent_key: str = '', sep: str = '_') -> Dict[str, Any]:
        """Flatten nested dictionary."""
        items = []

        for k, v in d.items():
            # Skip special keys
            if k.startswith('_'):
                continue

            new_key = f"{parent_key}{sep}{k}" if parent_key else k

            if isinstance(v, dict):
                items.extend(self._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # For lists, just count items or join strings
                if v and isinstance(v[0], str):
                    items.append((new_key, ', '.join(v)))
                else:
                    items.append((f"{new_key}_count", len(v)))
            else:
                items.append((new_key, v))

        return dict(items)

    def _get_default_css(self) -> str:
        """Get default CSS for HTML export."""
        return """
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        h1 {
            color: #333;
            border-bottom: 2px solid #4CAF50;
            padding-bottom: 10px;
        }
        .data-table {
            border-collapse: collapse;
            width: 100%;
            background-color: white;
            box-shadow: 0 1px 3px rgba(0,0,0,0.2);
        }
        .data-table th, .data-table td {
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }
        .data-table tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .data-table tr:hover {
            background-color: #f5f5f5;
        }
        .data-table th {
            background-color: #4CAF50;
            color: white;
            font-weight: bold;
            position: sticky;
            top: 0;
        }
        """


class ReportExporter:
    """Export reports in various formats with formatting."""

    def __init__(self):
        self.exporter = DataExporter()

    def export_employee_schedule(self, schedule_data: List[Dict[str, Any]],
                                 employee_name: str,
                                 format: str = ExportFormat.EXCEL) -> Union[str, bytes]:
        """Export employee schedule with special formatting."""
        if format == ExportFormat.EXCEL and EXCEL_AVAILABLE:
            return self._export_schedule_excel(schedule_data, employee_name)
        else:
            # Fall back to standard export
            return self.exporter.export(schedule_data, format)

    def _export_schedule_excel(self, schedule_data: List[Dict[str, Any]],
                               employee_name: str) -> bytes:
        """Create formatted Excel schedule."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Schedule for {employee_name}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Date", "Shift", "Start-End", "Location", "Notes"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        row_idx = 4
        current_week = None

        for entry in schedule_data:
            # Add week separator
            entry_date = entry.get('date')
            if entry_date:
                week_num = entry_date.isocalendar()[1]
                if week_num != current_week:
                    current_week = week_num
                    ws.merge_cells(f'A{row_idx}:E{row_idx}')
                    week_cell = ws.cell(row=row_idx, column=1)
                    week_cell.value = f"Week {week_num}"
                    week_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    week_cell.font = Font(bold=True)
                    row_idx += 1

            # Add schedule entry
            shift_info = entry.get('5SHIFT_related', {})
            workplace_info = entry.get('5WOPL_related', {})

            ws.cell(row=row_idx, column=1, value=str(entry.get('date', '')))
            ws.cell(row=row_idx, column=2, value=shift_info.get('name', ''))
            ws.cell(row=row_idx, column=3, value=shift_info.get('startend', ''))
            ws.cell(row=row_idx, column=4, value=workplace_info.get('name', ''))
            ws.cell(row=row_idx, column=5, value="")

            # Color weekends
            if entry_date and entry_date.weekday() >= 5:  # Saturday or Sunday
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
                    )

            row_idx += 1

        # Auto-size columns
        for col_idx in range(1, 6):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = [10, 20, 15, 20, 30][col_idx - 1]

        # Save
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_summary_report(self, report_data: Dict[str, Any],
                              format: str = ExportFormat.HTML) -> Union[str, bytes]:
        """Export a summary report with sections."""
        if format == ExportFormat.HTML:
            return self._export_summary_html(report_data)
        elif format == ExportFormat.MARKDOWN:
            return self._export_summary_markdown(report_data)
        else:
            # For other formats, flatten the data
            return self.exporter.export(report_data, format)

    def _export_summary_html(self, report_data: Dict[str, Any]) -> str:
        """Create HTML summary report."""
        html_parts = []

        # HTML header
        html_parts.append('<!DOCTYPE html>')
        html_parts.append('<html>')
        html_parts.append('<head>')
        html_parts.append(f'<title>{report_data.get("title", "Report")}</title>')
        html_parts.append('<meta charset="utf-8">')
        html_parts.append('<style>')
        html_parts.append(self._get_report_css())
        html_parts.append('</style>')
        html_parts.append('</head>')
        html_parts.append('<body>')

        # Report content
        html_parts.append(f'<h1>{html.escape(report_data.get("title", "Report"))}</h1>')

        if "generated_at" in report_data:
            html_parts.append(f'<p class="metadata">Generated: {report_data["generated_at"]}</p>')

        # Process sections
        for key, value in report_data.items():
            if key in ["title", "generated_at", "metadata"]:
                continue

            html_parts.append(f'<h2>{html.escape(key.replace("_", " ").title())}</h2>')

            if isinstance(value, dict):
                html_parts.append('<div class="section">')
                html_parts.append(self._dict_to_html_table(value))
                html_parts.append('</div>')
            elif isinstance(value, list):
                html_parts.append('<div class="section">')
                if value and isinstance(value[0], dict):
                    html_parts.append(self._list_to_html_table(value))
                else:
                    html_parts.append('<ul>')
                    for item in value:
                        html_parts.append(f'<li>{html.escape(str(item))}</li>')
                    html_parts.append('</ul>')
                html_parts.append('</div>')
            else:
                html_parts.append(f'<p>{html.escape(str(value))}</p>')

        html_parts.append('</body>')
        html_parts.append('</html>')

        return '\n'.join(html_parts)

    def _dict_to_html_table(self, data: Dict[str, Any]) -> str:
        """Convert dictionary to HTML key-value table."""
        html = '<table class="summary-table">'
        for key, value in data.items():
            html += '<tr>'
            html += f'<td class="key">{html.escape(key.replace("_", " ").title())}</td>'
            html += f'<td class="value">{html.escape(str(value))}</td>'
            html += '</tr>'
        html += '</table>'
        return html

    def _list_to_html_table(self, data: List[Dict[str, Any]]) -> str:
        """Convert list of dicts to HTML table."""
        if not data:
            return '<p>No data</p>'

        fields = list(data[0].keys())

        html = '<table class="data-table">'

        # Headers
        html += '<thead><tr>'
        for field in fields:
            html += f'<th>{html.escape(field.replace("_", " ").title())}</th>'
        html += '</tr></thead>'

        # Data
        html += '<tbody>'
        for record in data:
            html += '<tr>'
            for field in fields:
                value = record.get(field, "")
                html += f'<td>{html.escape(str(value))}</td>'
            html += '</tr>'
        html += '</tbody>'

        html += '</table>'
        return html

    def _export_summary_markdown(self, report_data: Dict[str, Any]) -> str:
        """Create Markdown summary report."""
        lines = []

        # Title
        lines.append(f"# {report_data.get('title', 'Report')}")
        lines.append("")

        if "generated_at" in report_data:
            lines.append(f"*Generated: {report_data['generated_at']}*")
            lines.append("")

        # Process sections
        for key, value in report_data.items():
            if key in ["title", "generated_at", "metadata"]:
                continue

            lines.append(f"## {key.replace('_', ' ').title()}")
            lines.append("")

            if isinstance(value, dict):
                for k, v in value.items():
                    lines.append(f"- **{k.replace('_', ' ').title()}**: {v}")
                lines.append("")
            elif isinstance(value, list):
                if value and isinstance(value[0], dict):
                    # Table
                    fields = list(value[0].keys())
                    lines.append("| " + " | ".join(f.replace('_', ' ').title() for f in fields) + " |")
                    lines.append("| " + " | ".join(["-" * 10 for _ in fields]) + " |")
                    for record in value:
                        values = [str(record.get(f, "")) for f in fields]
                        lines.append("| " + " | ".join(values) + " |")
                else:
                    # List
                    for item in value:
                        lines.append(f"- {item}")
                lines.append("")
            else:
                lines.append(str(value))
                lines.append("")

        return '\n'.join(lines)

    def _get_report_css(self) -> str:
        """Get CSS for report formatting."""
        return """
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            margin: 40px auto;
            max-width: 1200px;
            background-color: #f8f9fa;
            color: #333;
        }
        h1 {
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
            margin-bottom: 30px;
        }
        h2 {
            color: #34495e;
            margin-top: 30px;
            margin-bottom: 15px;
        }
        .metadata {
            color: #7f8c8d;
            font-style: italic;
            margin-bottom: 20px;
        }
        .section {
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .summary-table {
            width: 100%;
            border-collapse: collapse;
        }
        .summary-table td {
            padding: 10px;
            border-bottom: 1px solid #ecf0f1;
        }
        .summary-table .key {
            font-weight: bold;
            width: 30%;
            color: #2c3e50;
        }
        .data-table {
            width: 100%;
            border-collapse: collapse;
        }
        .data-table th, .data-table td {
            padding: 10px;
            text-align: left;
            border-bottom: 1px solid #ecf0f1;
        }
        .data-table th {
            background-color: #3498db;
            color: white;
            font-weight: bold;
        }
        .data-table tr:hover {
            background-color: #ecf0f1;
        }
        """